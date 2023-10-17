from dotenv import load_dotenv
from datetime import datetime
import openpyxl
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


def tipooperacao_str(nr_operacao):
    operacao_caixa = {
        1: 'Aplicação',
        2: 'Resgate Parcial',
        5: 'Resgate Total',
    }
    return operacao_caixa.get(nr_operacao)


def get_corpo_email(tipo_operacao):
    saudacao = "Bom dia" if datetime.now().hour < 12 else "Boa tarde"
    operacao = ""
    if tipo_operacao == 1:
        operacao = "aplicação"
    elif tipo_operacao in [2, 5]:
        operacao = "resgate"
    else:
        print('ERROR - no envio do tipo de operação')
        sys.exit()
    return f"{saudacao}!\nFavor realizar a {operacao} em anexo.\nObrigado,\n\nGabriel Assunção\n(031)99276-0244"


def enviar_email(tipo_operacao, caminho_novo_arquivo):
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = os.getenv('SMTP_PORT')
    remetente_email = os.getenv('REMETENTE_EMAIL')
    remetente_senha = os.getenv('REMETENTE_SENHA')
    destinatarios = ['gabriel@cartor.com.br']
    emails_cc = ['gabriel@abstratinvest.com']

    msg = MIMEMultipart()
    msg['From'] = remetente_email
    msg['To'] = ', '.join(destinatarios)
    msg['Subject'] = f"{tipooperacao_str(tipo_operacao)} - Vanguard II - fundo Soberano Santander"

    with open(caminho_novo_arquivo, 'rb') as arquivo:
        parte_anexada = MIMEBase('application', 'octet-stream')
        parte_anexada.set_payload(arquivo.read())
        encoders.encode_base64(parte_anexada)
        parte_anexada.add_header('Content-Disposition', 'attachment', filename=os.path.basename(caminho_novo_arquivo))
        msg.attach(parte_anexada)

    msg.attach(MIMEText(get_corpo_email(tipo_operacao), 'plain'))

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(remetente_email, remetente_senha)
        server.sendmail(remetente_email, destinatarios, msg.as_string())


def main():
    data_atual = datetime.now()
    data_formatada = data_atual.strftime('%d/%m/%y')
    tipo_operacao = int(input("\nTipo de Operação: "))
    valor_operacao = float(input("Valor da Operação: "))
    print('')

    if tipooperacao_str(tipo_operacao) is None:
        print('Tipo de Operação não existente. Favor Verificar!\n')
        sys.exit()

    load_dotenv()
    caminho_arquivo = os.getenv('CAMINHO_ARQUIVOCAIXA')
    caminho_template = os.path.join(caminho_arquivo, 'modelo_santander.xlsx')

    with open(caminho_template, 'rb') as arquivo_template:
        arquivo = openpyxl.load_workbook(arquivo_template)
        planilha = arquivo['NOVO MODELO SANTANDER']
        planilha.cell(5, 2).value = tipo_operacao
        planilha.cell(5, 6).value = valor_operacao
        planilha.cell(5, 8).value = data_formatada

        caminho_novo_arquivo = os.path.join(caminho_arquivo, f'VanguardII_{valor_operacao}.xlsx')
        arquivo.save(caminho_novo_arquivo)

    enviar_email(tipo_operacao, caminho_novo_arquivo)


if __name__ == "__main__":
    main()

