# script para movimentar caixa aplicações Santander
# caminho para o arquivo no sistema + abrir o arquivo em excel

from dotenv import load_dotenv
from datetime import datetime

import openpyxl
import os
import sys

import smtplib

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


def tipooperacao_str(nr_operacao):
    operacao_caixa = {
        1: 'Aplicação',
        2: 'Resgate Parcial',
        5: 'Resgate Total',    }
    operation_type = operacao_caixa.get(nr_operacao)
    return operation_type

data_atual = datetime.now()
data_formatada = data_atual.strftime('%d/%m/%y')
tipo_operacao = int(input("\nTipo de Operação: "))
valor_operacao = float(input("Valor da Operação: "))

if tipooperacao_str(tipo_operacao) is None:
    print('Tipo de Operação não existente. Favor Verificar!\n')
    sys.exit()

load_dotenv()
caminho_arquivo = os.getenv('CAMINHO_ARQUIVOCAIXA')
caminho_template = caminho_arquivo + '/modelo_santander.xlsx'
arquivo = openpyxl.load_workbook(caminho_template)


planilha = arquivo['NOVO MODELO SANTANDER']
planilha.cell(5, 2).value = tipo_operacao
planilha.cell(5, 6).value = valor_operacao
planilha.cell(5, 8).value = data_formatada


print('')
print(planilha.cell(5, 2).value)
print(planilha.cell(5, 6).value)
print(planilha.cell(5, 8).value)
print('')

caminho_novo_arquivo = caminho_arquivo+'/VanguardII_'+str(valor_operacao)+'.xlsx'
arquivo.save(caminho_novo_arquivo)

print(tipooperacao_str(tipo_operacao))

smtp_server = os.getenv('SMTP_SERVER')
smtp_port = os.getenv('SMTP_PORT')
remetente_email = os.getenv('REMETENTE_EMAIL')
remetente_senha = os.getenv('REMETENTE_SENHA')
destinatarios = ['gabriel@cartor.com.br']
msg = MIMEMultipart()
msg['From'] = remetente_email
msg['To'] = ', '.join(destinatarios)
msg['Subject'] = tipooperacao_str(tipo_operacao)+' - Vanguard II - fundo Soberano Santander'

nome_arquivo = caminho_novo_arquivo
parte_anexada = MIMEBase('application', 'octet-stream')
parte_anexada.set_payload(open(nome_arquivo, 'rb').read())
encoders.encode_base64(parte_anexada)
parte_anexada.add_header('Content-Disposition', 'attachment', filename=nome_arquivo.split('/')[-1])
msg.attach(parte_anexada)

if datetime.now().hour < 12:
    if tipo_operacao == 1:
        corpo_email = 'Bom dia!\nFavor realizar a aplicação em anexo.\nObrigado,\n\nGabriel Assunção\n(031)99276-0244' 
    elif tipo_operacao == 2 or tipo_operacao == 5:
        corpo_email = 'Bom dia!\nFavor realizar o resgate em anexo.\nObrigado,\n\nGabriel Assunção\n(031)99276-0244' 
    else:
        print('ERROR - no envio do tipo de operação')
        sys.exit()


else:
    if tipo_operacao == 1:
        corpo_email = 'Boa tarde!\nFavor realizar a aplicação em anexo.\nObrigado,\n\nGabriel Assunção\n(031)99276-0244' 
    elif tipo_operacao == 2 or tipo_operacao == 5:
        corpo_email = 'Boa tarde!\nFavor realizar o resgate em anexo.\nObrigado,\n\nGabriel Assunção\n(031)99276-0244'
    else:
        print('ERROR - no envio do tipo de operação')
        sys.exit()        

print(corpo_email)
# msg.attach(MIMEText(corpo_email, 'plain'))

# server = smtplib.SMTP(smtp_server, smtp_port)
# server.starttls()
# server.login(remetente_email, remetente_senha)
# server.sendmail(remetente_email, destinatarios, msg.as_string())
# tipo de operação = 1:Aplicaçãp., 2:Resgate Parcial, 5: Resgate Total

